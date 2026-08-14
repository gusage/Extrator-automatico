import pdfplumber
import re
import os
from datetime import datetime, timedelta, time as dtime
import openpyxl

# Regex para cabeçalho
padrao_funcionario = re.compile(r"(\d{6})\s*-\s*([A-ZÀ-ÚÇ\s]+?)\s+Admiss")
padrao_cargo = re.compile(r"Cargo:\s*\d+\s*-\s*([A-ZÀ-ÚÇ\s\-]+?)\s+Filial")
padrao_periodo = re.compile(r"Período\s*:\s*(\d{2}/\d{2}/\d{4})\s*a\s*(\d{2}/\d{2}/\d{4})")

# Faixas de coordenada X das colunas
X_DT_MIN, X_DT_MAX = 10, 32
X_PONTO_MIN, X_PONTO_MAX = 78, 230
X_EXTRA_MIN, X_EXTRA_MAX = 231, 530

padrao_hora = re.compile(r"^([0-2]?\d:[0-5]\d)$")

# Função para agrupar palavras por linha com base na coordenada Y
def agrupar_por_linha(palavras, tolerancia=2):
    linhas = []
    for p in sorted(palavras, key=lambda w: w["top"]):
        adicionado = False
        for linha in linhas:
            if abs(linha[0]["top"] - p["top"]) <= tolerancia:
                linha.append(p)
                adicionado = True
                break
        if not adicionado:
            linhas.append([p])
    return linhas

# Função para converter horário em minutos
def horario_para_minutos(h):
    horas, minutos = h.split(":")
    return int(horas) * 60 + int(minutos)

def remover_duplicados(horarios):
    vistos = []
    for h in horarios:
        if h not in vistos:
            vistos.append(h)
    return vistos

# Função para extrair os dias e horários de uma página do PDF
def extrair_dias(pagina):
    palavras = pagina.extract_words()
    linhas = agrupar_por_linha(palavras)

    dias = {}
    dia_atual = None

    for linha in linhas:
        linha_ordenada = sorted(linha, key=lambda w: w["x0"])

        if any(w["text"] == "Banco" for w in linha_ordenada):
            break

        possui_dt = any(
            X_DT_MIN <= w["x0"] <= X_DT_MAX and w["text"].isdigit() and 1 <= int(w["text"]) <= 31
            for w in linha_ordenada
        )

        if possui_dt:
            dt = next(w["text"] for w in linha_ordenada if X_DT_MIN <= w["x0"] <= X_DT_MAX)
            dia_atual = dt
            dias[dia_atual] = {"ponto": [], "intervalos": []}

        if dia_atual is None:
            continue

        for w in linha_ordenada:
            if not padrao_hora.match(w["text"]):
                continue
            if X_PONTO_MIN <= w["x0"] <= X_PONTO_MAX:
                dias[dia_atual]["ponto"].append(w["text"])
            elif X_EXTRA_MIN <= w["x0"] <= X_EXTRA_MAX:
                dias[dia_atual]["intervalos"].append(w["text"])

    return dias

#Função para escolher o intervalo mais adequado com base na entrada e na tolerância
def escolher_intervalo(entrada, intervalos, tolerancia=10):
    """
    Retorna o par (saída_intervalo, retorno_intervalo) mais adequado.
    Pula o 1º par se ele coincidir com a entrada (dentro da tolerância).
    """
    pares = []
    for i in range(0, len(intervalos) -1, 2):
        pares.append((intervalos[i], intervalos[i + 1]))

    if not pares:
        return None, None

    entrada_min = horario_para_minutos(entrada)

    for saida, retorno in pares:
        saida_min = horario_para_minutos(saida)
        if abs(saida_min - entrada_min) > tolerancia:
            return saida, retorno

    #Se todos os pares coincidirem, retorna o último
    return pares[-1]

#Função para calcular total de intervalos em minutos, considerando que o retorno pode ser no dia seguinte
def calcular_total_intervalos(saida, retorno):
    s = horario_para_minutos(saida)
    r = horario_para_minutos(retorno)
    total = r - s
    if total < 0:
        total += 24 * 60
    horas = total // 60
    minutos = total % 60
    return dtime(horas, minutos)

# Função para converter horário em objeto time
def horario_para_time(txt):
    if not txt:
        return None
    h, m = txt.split(":")
    return dtime(int(h), int(m))

#Função para gerar o arquivo Excel de um funcionário com base nos dados extraídos
def gerar_excel_funcionario(matricula, nome, data_inicio, ultimo_dia, dias_dict, pasta_saida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartao Ponto"
    ws.column_dimensions['A'].width = 12

    for d in range(1, ultimo_dia + 1):
        chave = f"{d:02d}"
        horarios = dias_dict.get(chave, [])

        data_atual = data_inicio + timedelta(days=d - 1)
        linha = d

        cel_data = ws.cell(row=linha, column=1, value=data_atual)
        cel_data.number_format = "dd/mm/yyyy"

        for j, h in enumerate(horarios):
            cel = ws.cell(row=linha, column=2 + j, value=horario_para_time(h))
            cel.number_format = "[h]:mm"

    nome_arquivo = f"{matricula} - {nome}.xlsx"
    caminho = os.path.join(pasta_saida, nome_arquivo)
    wb.save(caminho)
    print(f"  Gerado: {caminho}")

#Função para processar todos os PDFs em uma pasta e gerar arquivos Excel consolidados por funcionário
def processar_pasta(pasta_pdfs, pasta_saida, callback=None):
    os.makedirs(pasta_saida, exist_ok=True)

    pdfs = sorted([
        os.path.join(pasta_pdfs, f)
        for f in os.listdir(pasta_pdfs)
        if f.endswith(".pdf")
    ])

    # Dicionário acumulador por funcionário
    funcionarios = {}

    total_pdfs = len(pdfs)
    for idx_pdf, caminho_pdf in enumerate(pdfs):
        with pdfplumber.open(caminho_pdf) as pdf:
            paginas = [
                (i, p) for i, p in enumerate(pdf.pages)
                if p.extract_text() and "Marcações Ponto" in p.extract_text()
            ]
            for i, pagina in paginas:
                texto = pagina.extract_text()

                match_func = padrao_funcionario.search(texto)
                match_periodo = padrao_periodo.search(texto)

                if not match_func or not match_periodo:
                    continue

                matricula = match_func.group(1)
                nome = match_func.group(2).strip()
                data_inicio = datetime.strptime(match_periodo.group(1), "%d/%m/%Y")

                # Inicializa funcionário se ainda não existe
                if matricula not in funcionarios:
                    funcionarios[matricula] = {"nome": nome, "dias": {}}

                # Extrai dias e acumula
                dias = extrair_dias(pagina)
                for dia, grupos in dias.items():
                    data_atual = data_inicio + timedelta(days=int(dia) - 1)
                    chave = data_atual.strftime("%Y-%m-%d")
                    funcionarios[matricula]["dias"][chave] = grupos

        if callback:
            callback(idx_pdf + 1, total_pdfs, os.path.basename(caminho_pdf))

    # Gera um Excel por funcionário com todos os meses
    for matricula, dados in funcionarios.items():
        gerar_arquivos_funcionario(matricula, dados["nome"], dados["dias"], pasta_saida)

def gerar_arquivos_funcionario(matricula, nome, dias_dict, pasta_saida):
    #Cria subpasta do funcionario
    nome_pasta = f"{matricula} - {nome}"
    caminho_pasta = os.path.join(pasta_saida, nome_pasta)
    os.makedirs(caminho_pasta, exist_ok=True)

    chaves_ordenadas = sorted(dias_dict.keys())

    # Arquivos Horas
    wb_horas = openpyxl.Workbook()
    ws_horas = wb_horas.active
    ws_horas.title = "Horas"
    ws_horas.column_dimensions['A'].width = 12

    # Arquivos Intervalos
    wb_int = openpyxl.Workbook()
    ws_int = wb_int.active
    ws_int.title = "Intervalos"
    ws_int.column_dimensions['A'].width = 12

    for idx, chave in enumerate(chaves_ordenadas):
        grupos = dias_dict[chave]
        ponto = grupos["ponto"]
        intervalos = grupos["intervalos"]
        data = datetime.strptime(chave, "%Y-%m-%d")
        linha = idx + 1

        # Linha de arquivos Horas
        cel = ws_horas.cell(row=linha, column=1, value=data)
        cel.number_format = "dd/mm/yyyy"

        if ponto:
            entrada = ponto[0]
            saida_final = ponto[-1]

            saida_int, retorno_int = escolher_intervalo(entrada, intervalos)

            # Monta sequência de horários para o arquivo Horas
            sequencia_horas = [entrada]

            if saida_int and retorno_int:
                sequencia_horas.append(saida_int)
                sequencia_horas.append(retorno_int)

            sequencia_horas.append(saida_final)

            for j, h in enumerate(sequencia_horas):
                cel = ws_horas.cell(row=linha, column=2 + j, value=horario_para_time(h))
                cel.number_format = "[h]:mm"

        # Linha arquivos Intervalos
        cel_int = ws_int.cell(row=linha, column=1, value=data)
        cel_int.number_format = "dd/mm/yyyy"

        pares = []
        for i in range(0, len(intervalos) - 1, 2):
            pares.append((intervalos[i], intervalos[i + 1]))

        total_minutos = 0
        for j, (saida, retorno) in enumerate(pares):
            col_base = 2 + j * 2
            ws_int.cell(row=linha, column=col_base, value=horario_para_time(saida)).number_format = "[h]:mm"
            ws_int.cell(row=linha, column=col_base + 1, value=horario_para_time(retorno)).number_format = "[h]:mm"

            # Acumula o total sem escrever ainda
            s = horario_para_minutos(saida)
            r = horario_para_minutos(retorno)
            diff = r - s
            if diff < 0:
                diff += 24 * 60
            total_minutos += diff

        # Escreve o total geral na última coluna
        if pares:
            col_total = 2 + len(pares) * 2
            horas = total_minutos // 60
            minutos = total_minutos % 60
            cel_total = ws_int.cell(row=linha, column=col_total, value=dtime(horas, minutos))
            cel_total.number_format = "[h]:mm"
            cel_total.font = openpyxl.styles.Font(bold=True)

    # Salva os dois arquivos
    wb_horas.save(os.path.join(caminho_pasta, f"{nome_pasta} - Horas.xlsx"))
    wb_int.save(os.path.join(caminho_pasta, f"{nome_pasta} - Intervalos.xlsx"))
    print(f" Gerado: {caminho_pasta}")
    